import yfinance as yf
import os

import matplotlib.pyplot as plt
import pandas as pd
import math
import mpl_finance
import numpy as np
import matplotlib.dates as mdates

from mpl_finance import candlestick_ohlc
from datetime import date
import datetime

from ReadList import xls_handling
from data_prep import stockHandling



from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

greenFill = PatternFill(start_color='0099CC00',
                   end_color='0099CC00',
                   fill_type='solid')


# creates rolling means acoring the list sind back string for positive or neagtive trend
def fastAnalyse(series):
    """
    makes a string for gradients over means  2,5,25,96,200,500
    rerturns the string
    
    """
    
    avrg=[2,5,25,96,200,500]
    resultStr=""
    for i in avrg:
        mean_gradient=series.rolling(window=i).mean().diff()[-1]
        #print(mean_gradient)
        if mean_gradient > 0 :
            resultStr=resultStr + "+"
        else :
            resultStr=resultStr+"-"
    return resultStr


def get_Data_yahoo(ticker):
    global failList
    global i
    newDF=pd.DataFrame()
    tickerSymbol=ticker
    try:
        msft = yf.Ticker(tickerSymbol)
        stockinfo=msft.info
        stockName=str(stockinfo.get("longName"))  
    except:
        stockName=str(mySupervisionList[i]).split(".")[0]
        print("failed Ticker",mySupervisionList[i])
    try:
        newDF = yf.download(tickerSymbol, start="2020-02-01", end=today)    
    except:
        failList.append(mySupervisionList[i])
        print("failed download",mySupervisionList[i])
    print("stockName")    
    return stockName, newDF   







def find_trigger(kurs,fixedLimits,LimitFromDays): 
    """ 
    function should check if a trigger condition is met in the actual run
    return for found trigger false for no trigger
"""
    foundTriggerList=[] 
    # hier all limit nur ein limit abfangen count=0
    kurs=kurs.iloc[-20:]
    for each in fixedLimits:
        mdf["high_treshold"]=(kurs["close"]>each)       
        mdf["high_treshold_shift"]=mdf["high_treshold"].shift(periods=1)    
        mdf["Trigger"]=mdf["high_treshold"] != mdf["high_treshold_shift"]
        triggerArray=np.array(mdf["Trigger"])
        # find oslution for [0] outer bound
        foundTriggerList.append(np.argwhere(triggerArray[1:]==True)[0][0]) 
        
        
    for each in LimitFromDays:
        mdf["high_treshold"]=(kurs["close"]>each)       
        mdf["high_treshold_shift"]=mdf["high_treshold"].shift(periods=1)    
        mdf["Trigger"]=mdf["high_treshold"] != mdf["high_treshold_shift"]
        triggerArray=np.array(mdf["Trigger"])
        foundTriggerList.append(np.argwhere(triggerArray[1:]==True)) 
    if len(foundTriggerList)>0:
        return True
    
    return False




def get_my_list():
    """
    # replaced by xlsx
    
    #generate a list of stock for supervision
    #if file is available or take the stored data here
    """
    try:
        returnList=pd.read_csv("TestList.csv",header= None)
        print("importetd List",returnList)
        returnList=returnList.iloc[0].to_list()
    
    except:    
        watchlist=["TMV.DE","AAPL","HEIA.AS","SHL.DE","VOS.DE","GE","SKT","JNJ","AVGO","SW1.F","SLT.DE","ASL.de"]
        mySupervisionList=["2338.HK","AAG.DE","BIDU","BMW.DE","BAYN.DE","COK.DE","CSCO","EVD.DE","FEV.DE","HAG.F","IRBT","JD","MTX.DE","N7G.DE","PRLB","SHL.DE","SIX2.DE","SLM","TCOM","TUI1.DE","VOW3.DE"]
        returnList=watchlist+mySupervisionList
        
    return returnList



def createTodaysFolders():
    today = date.today()
    print("Today's date:", today)
    try:
        os.mkdir("save//pics//"+str(today))
        os.mkdir("save//files//"+str(today))
    except OSError:
        print ("Creation of the directory failed use existing" )
    else:
        print ("Successfully created the directory " )
    return today


def makeCandles(df):
    df_ohlc=df["Adj Close"].resample("2D").ohlc()
    df_ohlc_1d=df["Adj Close"].resample("1D").ohlc()
    df_Volume = df["Volume"].resample("2D").sum()
    df["100ma"]=df["Adj Close"].rolling(window=100).mean()
    df["20ma"]=df["Adj Close"].rolling(window=20).mean()
    df["df_dx"]=df["Adj Close"]-df["100ma"].rolling(window=50).mean()
    

    df_ohlc.reset_index(inplace=True)
    df_ohlc["Date"]= df_ohlc["Date"].map(mdates.date2num)
    df_ohlc_1d.reset_index(inplace=True)
    df_ohlc_1d["Date_time"]= df_ohlc_1d["Date"].map(mdates.date2num)
    
#    df_ohlc_1d.to_csv("save//files//"+str(today)+"//"+str(stockName)+".csv")

    return[df, df_ohlc,df_Volume,df_ohlc_1d]


def draw_limits(all_limits):
    
    ax1=plt.subplot2grid((6,1),(0,0),rowspan=5,colspan=1)
    for count in range(len(all_limits)):  
        # fixed limit line
        ax1.axhline(y=all_limits.loc[count,"fix Limit"], color='r', linestyle='-')    
        # dyn in day line
        ax1.axhline(y=all_limits.loc[count,"Last Days Limit"], color='g', linestyle='-')    

def make_log(name,value, typeOfTrigger):
    file1 = open("Mylogfile.txt","a")
    file1.write(str(name)+"\t\t"+str(value)+"\t"+str(typeOfTrigger)+"\n")
    file1.close()


def findLimits():
    fixedLimits=[]
    limitCols=["F","G"]
    for each in limitCols:
        threshold=np.float32(abc.ws[str(each)+str(i+2)].value)      
        stockPrice=np.float32(abc.ws["E"+str(i+2)].value)
        fixedLimits.append(threshold)
        # paint cell with limits
        if (threshold > stockPrice):
            print("check out {} level {}".format(stockName,threshold))
            abc.ws[str(each)+str(i+2)].fill = redFill
            make_log(stockName,threshold, "fixed Limit")
        if (threshold < stockPrice):
            abc.ws[str(each)+str(i+2)].fill = greenFill
    
    only_Limit_list = [x for x in fixedLimits if np.isnan(x) == False]
    return only_Limit_list

def checkSignals():
    """
    hier eine funktion die ma10,20 38 auf ositiv prüft und ein kauf signal generiert"
    df_ohlc_1d.iloc[-1,1] > ma10
    
    &
    
    """
    pass

def moving_alarm_value(xlsSheet,df):   
    """ nimt zellen datum und prozent für ein training stop limit
    dies wird in DF geschrieben
    """
    abc=xlsSheet
    start_date=abc.ws["I"+str(i+2)].value
    start_value=abc.ws["H"+str(i+2)].value
    if (start_date is not None):
        print("found dynmaik limit")
        try:
            trailingDF=df.loc[df.index > start_date ]
        except:
            print(i,start_date, start_date is datetime.date)
        limitList=[0]*(len(df)-len(trailingDF))
        kurList=[]

        # parsing history for adaptiing treshold for stop buy limit
        
        if start_value >0:
            limit_old=(((trailingDF.iloc[0].Close+trailingDF.iloc[0].Open)/2)*(1+(start_value/100)))
            print("stop buy limit {} a lat open {} last close {}".format(limit_old,trailingDF.iloc[0].Close,trailingDF.iloc[0].Open))
            for index, values in trailingDF.iterrows():
                try:
                    limit=(float((values.Open+values.Close))/2)*(1+(start_value/100))
                except:
                    pass
                kurList.append(limit)
                if limit<limit_old:
                    limit_old=limit
                limitList.append(limit_old)
            df["trailing_limit"]=limitList
            if (limit_old < df.iloc[-1].Open) :
                abc.ws[str("I")+str(i+2)].fill = redFill
            if (limit_old > df.iloc[-1].Open):
                abc.ws[str("I")+str(i+2)].fill = greenFill            
            
        #Parsing for treshold trailing stop limit
            
        if start_value <0:    
            limit_old=(((trailingDF.iloc[0].Close+trailingDF.iloc[0].Open)/2)*(1-(abs(start_value)/100)))
            print("trailing loss limit")

            for index, values in trailingDF.iterrows():
                try:
                    limit=(float((values.Open+values.Close))/2)*(1-(abs(start_value)/100))
                except:
                    pass
                kurList.append(limit)
                if limit>limit_old:
                    limit_old=limit
                limitList.append(limit_old)
            df["trailing_limit"]=limitList
            if (limit_old > df.iloc[-1].Open) :
                abc.ws[str("I")+str(i+2)].fill = redFill
                make_log(stockName,limit, "trailing Limit")
            if (limit_old < df.iloc[-1].Open):
                abc.ws[str("I")+str(i+2)].fill = greenFill            
    
    return
    
    


##main



failList=[]
RSL_List=[]
trendIndicatorList=[]
mdf=pd.DataFrame()
stockNameList=[]



### main
#import list to check
#mySupervisionList=get_my_list()


file="LimitCheck.xlsx"
#file="Test_limit.xlsx"


abc=xls_handling()

abc.load_list_from_xls(file,"TestList")
allSheets=abc.getSheets()

print("choose onen´of the sheets")
for eachsheet in range(len(allSheets)-1):
    print("nr {}.  foor sheet name {}".format(eachsheet,allSheets[eachsheet]))
    abc.load_list_from_xls(file,allSheets[eachsheet])

eachsheet=0
#sheeList=abc.sheetnames
backlook=200

mySupervisionList=abc.getKeys(0)


print("imported stocks from xls: ", mySupervisionList)

today=createTodaysFolders()

make_log("*********",str(today), "*******************\n")
    
for i in range(len(mySupervisionList)):   
    stockName,df = get_Data_yahoo(mySupervisionList[i])   
    print(stockName)
    abc.ws["B"+str(i+2)]=stockName
    
    
    # hier was rein if df is epmty pass passiert bei download und empty df
    if (len(df) <1):
        print("hier war ein Yahoo Download Fehler",mySupervisionList[i])
        
    else:
        df, df_ohlc,df_Volume,df_ohlc_1d=makeCandles(df)
        
        
        
        RSL=df["Adj Close"].rolling(window=3).mean()/df["Adj Close"].rolling(window=5*26).mean()
        RSL_List.append(RSL[-1])
        
        # fast analyse + - over averages 2,5,25,96,200,500
        trendIndicator=fastAnalyse(df["Adj Close"])
        trendIndicatorList.append(trendIndicator)
        stockNameList.append(str(stockName))
        # chang this to relative not absulute
        abc.ws["B"+str(i+2)]=stockName     #Full Name
        abc.ws["C"+str(i+2)]=trendIndicator#TrendIndikator
        abc.ws["D"+str(i+2)]=str(RSL[-1])[:4]#RSL
        # mark RSI in red and green
        if (float(abc.ws["D"+str(i+2)].value) < 0.9):
            abc.ws["D"+str(i+2)].fill = redFill
        if (float(abc.ws["D"+str(i+2)].value) >= 1.1):
            abc.ws["D"+str(i+2)].fill = greenFill
        
        abc.ws["E"+str(i+2)]=df_ohlc_1d.iloc[-1,1]#last value

        checkSignals()       # soll für signal trigger was werden 
        
        
        fixedLimits=findLimits()            # macht limit aus festen weerten aus dem xls
        moving_alarm_value(abc,df)          # macht ein trailing limit
        
        abc.wb.save("save//files//"+str(today)+"Result_"+str(allSheets[eachsheet])+str(today)+".xlsx")



        #start plottting
        # make limit lines    
        ax1=plt.subplot2grid((6,1),(0,0),rowspan=5,colspan=1)
        if len(fixedLimits)!=False:
            # draw_limits(all_limit_values)
            for count in fixedLimits:  
                # fixed limit line
                if count!= None :
                   ax1.axhline(y=count, color='r', linestyle='-') 
               
        # plot picture
        plt.title(str(stockName)+" RSL : "+str(RSL[-1])[:4]+trendIndicator)
        ax2=plt.subplot2grid((6,1),(5,0),rowspan=5,colspan=1,sharex=ax1)
        ax1.xaxis_date()
        candlestick_ohlc(ax1,df_ohlc.values,width=2,colorup="g")
        ax1.plot(df.index.map(mdates.date2num),df["100ma"],label="100ma")
        ax1.plot(df.index.map(mdates.date2num),df["20ma"],label="20ma")
        if "trailing_limit" in df.keys():
            ax1.plot(df.index.map(mdates.date2num),df["trailing_limit"],label='trailing_loss',c="green",linewidth=4)
        
        ax2.fill_between(df_Volume.index.map(mdates.date2num), df_Volume.values,0)
        ax1.annotate(str(df["Adj Close"][-1])[:6],xy=(0.8, 0.9), xycoords='axes fraction')
        ax2.plot(df.index.map(mdates.date2num),df["df_dx"]) 
        ax1.set_xlim([today-0.3*datetime.timedelta(backlook), today])
        ax2.set_xlim([today-0.3*datetime.timedelta(backlook), today])
        plt.savefig("save//pics//"+str(today)+"//"+str(stockName).split(".")[0]+str(today), dpi=800)
        ax1.set_xlim([today-datetime.timedelta(backlook), today])
        ax2.set_xlim([today-datetime.timedelta(backlook), today])
        #plt.legend()
        plt.show()



def plot_bollinger(DF):
    mdf=pd.DataFrame()
    mdf["value"]=DF
    mdf=mdf.dropna()
    mdf["max"]=mdf["value"].rolling(10).max()
    mdf["min"]=mdf["value"].rolling(10).min()
    mdf['TP'] = (mdf['value'] + mdf['max'] + mdf['min'])/3
    mdf['MA-TP'] = mdf['TP'].rolling(4).mean()
    mdf['stdw'] = mdf['TP'].rolling(10).std(ddof=1)
    mdf['mean_max'] = mdf['MA-TP']+ 4*mdf['stdw']
    mdf['mean_min'] = mdf['MA-TP'] - 4*mdf['stdw']
    
    # Plotting it all together
    ax = mdf[['value', 'mean_max', 'mean_min']].plot(color=['blue', 'orange', 'yellow'])
    ax.fill_between(mdf.index, mdf['mean_max'], mdf['mean_min'], facecolor='orange', alpha=0.3)
    ax.set_xlim([today-datetime.timedelta(70), today])
    plt.show()
    
    
#plot_bollinger(df["20ma"])


