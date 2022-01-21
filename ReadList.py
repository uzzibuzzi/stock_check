# -*- coding: utf-8 -*-
"""
Created on Wed Oct  6 07:51:42 2021

@author: vollmera
"""
import pandas as pd
import numpy as np

import yfinance as yf
import os

import openpyxl




from openpyxl import load_workbook


class xls_handling:
    def __init__(self):
        self.socklist = []
        self.wb=0
        self.ws=0
        self.stocklist=[]
        
    def load_list_from_xls(self,filename,sheetname):
        self.wb= load_workbook(filename)
        self.ws= self.wb[sheetname]
        wknRow=1
        
        for row in self.ws.iter_rows(min_row=2, max_col=wknRow, max_row=5):
            for cell in row:
                self.stocklist.append(cell.value)

    def getSheets(self):
        """return list of sheetnames"""
        return self.wb.sheetnames
    
    def getKeys(self,axis):
        newList=[]
        columEnd=1
        rowEnd=1
        if axis == 1:
            columEnd=100
            startRow=1
            #print("columEnd")
        if axis==0:
            rowEnd=100
            startRow=2
            #print("row end")
            
        for col in self.ws.iter_cols(min_row=startRow, max_col=columEnd, max_row=rowEnd):
            for cell in col:
                if cell.value is None:
                     return newList
                newList.append(cell.value)
        print("End of counter reached")
        return newList
  
    def get_price(self,symbol):
        # get the symbol out of the workbook
        tickerSymbol=symbol
        try:
            msft = yf.Ticker(tickerSymbol)
            stockinfo=msft.info
            stockinfo.get("longName")
            StockPrice=stockinfo.get("ask")
            return StockPrice
        except:
            return tickerSymbol
    def get_info(self,symbol,info):
        tickerSymbol=symbol

        try:
            msft = yf.Ticker(tickerSymbol)
            stockinfo=msft.info
            try :

                return stockinfo.get(str(info))
            except:

                return False
        except:
                return False                    
        
            
    def get_colum_From(self,label):
        aaa=self.getKeys(1)
        column=aaa.index(label)  
        rowList=[]
        for ROWS  in self.ws.iter_rows(1, len(abc.getKeys(0))):
            rowList.append(ROWS[column].value)
        return rowList
               
    def get_row_From(self,label):
        aaa=self.getKeys(0)
        rowselect=aaa.index(label)+1 
        ColList=[]
        for COLUM  in self.ws.iter_cols(1, len(abc.getKeys(1))):
            ColList.append(COLUM[rowselect].value)
        return ColList





if __name__ == "__main__":
    file="LimitCheck.xlsx"

    abc=xls_handling()
    abc.load_list_from_xls(file,"TestList")
    allSheets=abc.getSheets()
    allSheets.append("all")
    for each in allSheets:
        print(each)    

    
"""

    StockList=abc.getKeys(1)
    
    for each in StockList:
        print(abc.get_price(each))

    abc.ws['B4'] =  abc.get_info("TMV.DE","ask")
    abc.ws['B5'] =  abc.get_info("TMV.DE","longName")
        
    
    averageVolume=abc.get_info("SLM","averageVolume")
    averageVolume10days=abc.get_info("SLM","averageVolume10days")
    
    averageVolume10days/averageVolume
      
    abc.getKeys(1).index("Limit oben")  
    
    abc.get_colum_From("WKN")
    
    abc.get_row_From("BMW.DE")["Limit oben"]


"""

